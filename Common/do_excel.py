from openpyxl import load_workbook
from Common import dir_path
from Common.read_config import ReadConfig


class DoExcel:

    def __init__(self, option, section='archives.xlsx'):
        self.section = section
        self.option = option

    def get_data(self, file_path):  # 从excel中获取测试数据
        mode = eval(ReadConfig.get_config_data(dir_path.config_path, self.section, self.option))  # 读取配置文件，拿到配置文件中的字典
        test_datas = []
        for key in mode:      # 遍历字典中的key
            wb = load_workbook(file_path)
            sheet = wb[key]  # 获取到表单
            headers = []
            for i in range(1, sheet.max_column+1):   # 获取表中每一列的标题，添加到列表里面
                headers.append(sheet.cell(3, i).value)
            if mode[key] == "all":    # 如果配置文件中，表单名对应的值为all，就加载这个表单中所有的行
                for i in range(4, sheet.max_row+1):  # 从第二行开始，第一行是title
                    row_data = {}
                    for j in range(1, sheet.max_column+1):
                        row_data[headers[j-1]] = sheet.cell(i, j).value
                        row_data['sheet_name'] = key  # 在字典中加入表单名，方便在执行测试用例时调用，执行用例时，需将测试结果写回到对应的表单中
                    test_datas.append(row_data)   # 将字典添加到列表中
            else:
                for case_id in mode[key]:   # 配置文件中，表单名对应的值为一个列表，该字典为excel中用例数据对应的编号，遍历该字典，拿到用例编号
                    row_data = {}
                    for j in range(1, sheet.max_column+1):
                        row_data[headers[j-1]] = sheet.cell(case_id+3, j).value
                        row_data['sheet_name'] = key  # 在字典中加入表单名，方便在执行测试用例时调用，执行用例时，需将测试结果写回到对应的表单中
                    test_datas.append(row_data)   # 将字典添加到列表中
        return test_datas

    def write_back(self, file_path, sheet_name, i, j, value):  # 将数据写入excel
        wb = load_workbook(file_path)
        sheet = wb[sheet_name]
        sheet.cell(i, j).value = value   # i代表行，j代表列
        wb.save(file_path)


if __name__ == '__main__':
    # print(DoExcel().mode)
    a = DoExcel('data_point').get_data(dir_path.archives_data_excel_path)
    # print(len(a))
    print(a)
